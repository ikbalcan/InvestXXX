"""
Robot Portföy Yöneticisi Tab - Günlük Portföy Önerileri
"""

import streamlit as st
import pandas as pd
import numpy as np
import sys
import os
from datetime import datetime, timedelta
import json
import hashlib
import io

# Proje modüllerini import et
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))
sys.path.append(os.path.dirname(__file__))

from dashboard_utils import load_config, load_stock_data
from dashboard_stock_hunter import analyze_single_stock, train_model_for_symbol
from price_target_predictor import PriceTargetPredictor
from src.data_loader import DataLoader
from src.database import Database
from src.auth import require_auth, init_session_state

# Excel export için openpyxl kontrolü
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Portföy verilerini sakla
PORTFOLIO_FILE = 'logs/robot_portfolio.json'
TRANSACTIONS_DIR = 'logs/transactions'

def load_portfolio(user_id=None):
    """Portföy verilerini veritabanından yükle"""
    if user_id:
        db = get_db()
        return db.get_user_portfolio(user_id)
    
    # Fallback: Eski JSON dosyası (backward compatibility)
    if os.path.exists(PORTFOLIO_FILE):
        try:
            with open(PORTFOLIO_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    
    return {
        'cash': 0,
        'stocks': {}  # {symbol: {'quantity': int, 'avg_cost': float}}
    }

def save_portfolio(portfolio, user_id=None):
    """Portföy verilerini veritabanına kaydet"""
    if user_id:
        db = get_db()
        db.update_user_portfolio(user_id, portfolio)
    else:
        # Fallback: Eski JSON dosyası (backward compatibility)
        os.makedirs(os.path.dirname(PORTFOLIO_FILE), exist_ok=True)
        with open(PORTFOLIO_FILE, 'w', encoding='utf-8') as f:
            json.dump(portfolio, f, indent=2, ensure_ascii=False)

def format_currency(amount):
    """Parayı 2 ondalık basamakla formatla"""
    return round(amount, 2)

def get_all_bist_stocks():
    """Tüm BIST hisselerini döndür"""
    return [
        'THYAO.IS', 'AKBNK.IS', 'BIMAS.IS', 'EREGL.IS', 'FONET.IS', 'GARAN.IS',
        'ISCTR.IS', 'KRDMD.IS', 'PETKM.IS', 'SAHOL.IS', 'TUPRS.IS', 'ALBRK.IS',
        'ASELS.IS', 'FROTO.IS', 'HALKB.IS', 'TSKB.IS', 'VAKBN.IS', 'VAKFN.IS',
        'YKBNK.IS', 'CCOLA.IS', 'DOHOL.IS', 'ENKAI.IS', 'KCHOL.IS', 'KOZAL.IS',
        'MGROS.IS', 'OTKAR.IS', 'SISE.IS', 'TCELL.IS', 'TOASO.IS', 'TKFEN.IS',
        'ULKER.IS', 'VESTL.IS', 'ZOREN.IS', 'ARCLK.IS', 'AZTEK.IS', 'NETAS.IS',
        'PAMEL.IS', 'SELEC.IS', 'SMRTG.IS', 'TATGD.IS', 'ERSU.IS', 'KONYA.IS',
        'MARTI.IS', 'UNYEC.IS', 'GENIL.IS', 'PGSUS.IS', 'MEGMT.IS'
    ]

def get_user_id():
    """Kullanıcı kimliğini al - Veritabanından"""
    init_session_state()
    
    if 'authenticated' in st.session_state and st.session_state.authenticated:
        return st.session_state.user_id
    
    # Eğer authenticated değilse None döndür (auth gerekli)
    return None

def get_db():
    """Veritabanı instance'ını al"""
    if 'db' not in st.session_state:
        st.session_state.db = Database()
    return st.session_state.db

def load_user_transactions(user_id):
    """Kullanıcı işlemlerini veritabanından yükle"""
    if not user_id:
        return []
    
    db = get_db()
    transactions = db.get_user_transactions(user_id)
    
    # Formatı eski sisteme uyumlu hale getir
    formatted_transactions = []
    for t in transactions:
        formatted_transactions.append({
            'id': t['id'],
            'type': t['type'],
            'symbol': t['symbol'],
            'quantity': t['quantity'],
            'price': t['price'],
            'total_value': t['total_value'],
            'date': t['date'],
            'created_at': t['created_at']
        })
    
    return formatted_transactions

def save_user_transaction(user_id, transaction):
    """Kullanıcı işlemini veritabanına kaydet"""
    if not user_id:
        return None
    
    db = get_db()
    transaction_id = db.add_transaction(user_id, transaction)
    
    # Eski format ile uyumluluk için
    transaction['id'] = transaction_id
    transaction['created_at'] = datetime.now().isoformat()
    return transaction

def delete_user_transaction(user_id, transaction_id):
    """Kullanıcı işlemini veritabanından sil"""
    if not user_id:
        return False
    
    db = get_db()
    return db.delete_transaction(user_id, transaction_id)

def calculate_profit_loss(transactions):
    """İşlemlerden kar/zarar hesapla"""
    # FIFO mantığı ile kar/zarar hesapla
    buy_transactions = [dict(t) for t in transactions if t['type'] == 'AL']  # Kopya oluştur
    sell_transactions = [t for t in transactions if t['type'] == 'SAT']
    
    # Hisse bazında kar/zarar hesapla
    symbol_profits = {}
    
    for sell in sell_transactions:
        symbol = sell['symbol']
        sell_quantity = sell['quantity']
        sell_price = sell['price']
        sell_date = datetime.fromisoformat(sell['date'])
        
        # Bu satış için alış işlemlerini bul (FIFO)
        remaining_sell = sell_quantity
        total_cost = 0
        
        # Bu hisse için alış işlemlerini tarihe göre sırala
        symbol_buys = [b for b in buy_transactions if b['symbol'] == symbol]
        symbol_buys = sorted(symbol_buys, key=lambda x: x['date'])
        
        for buy in symbol_buys:
            if buy['quantity'] > 0:
                buy_date = datetime.fromisoformat(buy['date'])
                if buy_date <= sell_date:  # Satıştan önceki alışlar
                    if remaining_sell > 0:
                        used_quantity = min(remaining_sell, buy['quantity'])
                        total_cost += used_quantity * buy['price']
                        buy['quantity'] -= used_quantity  # Kullanılan miktarı düş
                        remaining_sell -= used_quantity
        
        # Kar/zarar hesapla
        sell_value = sell_quantity * sell_price
        profit_loss = sell_value - total_cost
        profit_loss_pct = (profit_loss / total_cost * 100) if total_cost > 0 else 0
        
        if symbol not in symbol_profits:
            symbol_profits[symbol] = {'profit_loss': 0, 'profit_loss_pct': 0, 'count': 0}
        
        symbol_profits[symbol]['profit_loss'] += profit_loss
        symbol_profits[symbol]['count'] += 1
    
    # Toplam kar/zarar
    total_profit_loss = sum([p['profit_loss'] for p in symbol_profits.values()])
    
    return {
        'symbol_profits': symbol_profits,
        'total_profit_loss': total_profit_loss
    }

def calculate_remaining_positions(transactions):
    """Alış-satış işlemlerinden kalan pozisyonları hesapla (FIFO mantığı ile)"""
    # Alış ve satış işlemlerini ayır
    buy_transactions = [dict(t) for t in transactions if t['type'] == 'AL']  # Kopya oluştur
    sell_transactions = sorted([t for t in transactions if t['type'] == 'SAT'], key=lambda x: x['date'])
    
    # Hisse bazında kalan pozisyonları hesapla
    remaining_positions = {}
    
    # Tüm alış işlemlerini hisse bazında grupla
    for buy in buy_transactions:
        symbol = buy['symbol']
        if symbol not in remaining_positions:
            remaining_positions[symbol] = {
                'transactions': [],
                'total_quantity': 0,
                'total_cost': 0
            }
        
        remaining_positions[symbol]['transactions'].append({
            'quantity': buy['quantity'],
            'price': buy['price'],
            'date': buy['date']
        })
        remaining_positions[symbol]['total_quantity'] += buy['quantity']
        remaining_positions[symbol]['total_cost'] += buy['quantity'] * buy['price']
    
    # Satış işlemlerini FIFO mantığı ile uygula
    for sell in sell_transactions:
        symbol = sell['symbol']
        sell_quantity = sell['quantity']
        sell_date = datetime.fromisoformat(sell['date'])
        
        if symbol in remaining_positions:
            # Bu hisse için alış işlemlerini tarihe göre sırala (FIFO)
            symbol_transactions = sorted(
                remaining_positions[symbol]['transactions'],
                key=lambda x: x['date']
            )
            
            remaining_sell = sell_quantity
            
            # FIFO mantığı ile satışları alışlardan düş
            for buy_tx in symbol_transactions:
                if remaining_sell <= 0:
                    break
                
                buy_date = datetime.fromisoformat(buy_tx['date'])
                if buy_date <= sell_date:  # Satıştan önceki alışlar
                    if buy_tx['quantity'] > 0:
                        used_quantity = min(remaining_sell, buy_tx['quantity'])
                        
                        # Kullanılan miktarı ve maliyeti düş
                        buy_tx['quantity'] -= used_quantity
                        remaining_positions[symbol]['total_quantity'] -= used_quantity
                        remaining_positions[symbol]['total_cost'] -= used_quantity * buy_tx['price']
                        
                        remaining_sell -= used_quantity
    
    # Kalan pozisyonları portföy formatına dönüştür
    portfolio_stocks = {}
    
    for symbol, data in remaining_positions.items():
        remaining_quantity = data['total_quantity']
        
        if remaining_quantity > 0:
            # Ortalama maliyet hesapla
            avg_cost = data['total_cost'] / remaining_quantity if remaining_quantity > 0 else 0
            
            portfolio_stocks[symbol] = {
                'quantity': int(remaining_quantity),
                'avg_cost': round(avg_cost, 2)
            }
    
    return portfolio_stocks

def export_transactions_to_excel(transactions, profit_loss_data):
    """İşlemleri Excel formatında export et"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl modülü yüklü değil. Lütfen şu komutu çalıştırın: pip install openpyxl"
        )
    
    # İşlemler DataFrame'i
    if transactions:
        df_transactions = pd.DataFrame(transactions)
        if not df_transactions.empty and 'date' in df_transactions.columns:
            df_transactions = df_transactions[['date', 'symbol', 'type', 'quantity', 'price', 'total_value']]
            df_transactions['date'] = pd.to_datetime(df_transactions['date']).dt.strftime('%d.%m.%Y')
            df_transactions.columns = ['Tarih', 'Hisse', 'İşlem Tipi', 'Adet', 'Birim Fiyat (TL)', 'Toplam Tutar (TL)']
        else:
            df_transactions = pd.DataFrame(columns=['Tarih', 'Hisse', 'İşlem Tipi', 'Adet', 'Birim Fiyat (TL)', 'Toplam Tutar (TL)'])
    else:
        df_transactions = pd.DataFrame(columns=['Tarih', 'Hisse', 'İşlem Tipi', 'Adet', 'Birim Fiyat (TL)', 'Toplam Tutar (TL)'])
    
    # Kar/Zarar DataFrame'i - Hisse bazında detaylar
    symbol_profits = profit_loss_data.get('symbol_profits', {})
    profit_data = []
    for symbol, data in symbol_profits.items():
        profit_data.append({
            'Hisse': symbol,
            'Toplam Kar/Zarar (TL)': data['profit_loss'],
            'Kar/Zarar (%)': data['profit_loss_pct'],
            'İşlem Sayısı': data['count']
        })
    
    df_profits = pd.DataFrame(profit_data)
    
    # Toplam metrikleri hesapla
    total_buy = sum([t['total_value'] for t in transactions if t['type'] == 'AL'])
    total_sell = sum([t['total_value'] for t in transactions if t['type'] == 'SAT'])
    total_profit_loss = profit_loss_data.get('total_profit_loss', 0)
    total_transactions = len(transactions)
    total_buy_count = len([t for t in transactions if t['type'] == 'AL'])
    total_sell_count = len([t for t in transactions if t['type'] == 'SAT'])
    
    # Excel dosyası oluştur
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. İşlemler sayfası
        df_transactions.to_excel(writer, sheet_name='İşlemler', index=False)
        
        # 2. Kar/Zarar Analizi sayfası
        # Özet metrikleri DataFrame olarak oluştur
        summary_df = pd.DataFrame({
            'Metrik': [
                'Toplam Alış',
                'Toplam Satış',
                'Toplam Kar/Zarar',
                'Toplam İşlem Sayısı',
                'Alış İşlem Sayısı',
                'Satış İşlem Sayısı'
            ],
            'Değer': [
                f"{total_buy:,.2f} TL",
                f"{total_sell:,.2f} TL",
                f"{total_profit_loss:+,.2f} TL",
                total_transactions,
                total_buy_count,
                total_sell_count
            ]
        })
        
        if OPENPYXL_AVAILABLE:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Önce özet metrikleri yaz (satır 2'den başla, başlık için yer bırak)
            summary_df.to_excel(writer, sheet_name='Kar/Zarar Analizi', index=False, startrow=2)
            
            # Hisse bazında detayları yaz (özet metriklerden sonra)
            start_row = len(summary_df) + 5  # Başlık(1) + Özet başlığı(1) + Özet(6) + boş(1) = 9. satır
            if not df_profits.empty:
                df_profits.to_excel(writer, sheet_name='Kar/Zarar Analizi', index=False, startrow=start_row)
            else:
                # Boş DataFrame oluştur
                empty_df = pd.DataFrame(columns=['Hisse', 'Toplam Kar/Zarar (TL)', 'Kar/Zarar (%)', 'İşlem Sayısı'])
                empty_df.to_excel(writer, sheet_name='Kar/Zarar Analizi', index=False, startrow=start_row)
            
            # Workbook ve worksheet'i al
            workbook = writer.book
            worksheet = writer.sheets['Kar/Zarar Analizi']
            
            # Ana başlık ekle
            worksheet.merge_cells('A1:B1')
            title_cell = worksheet['A1']
            title_cell.value = '💰 Kar/Zarar Analizi Özeti'
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Özet başlığı
            worksheet['A2'] = 'Özet Metrikler'
            worksheet['A2'].font = Font(bold=True, size=12)
            
            # Hisse bazında başlık
            if not df_profits.empty:
                worksheet[f'A{start_row}'] = 'Hisse Bazında Kar/Zarar Detayları'
                worksheet[f'A{start_row}'].font = Font(bold=True, size=12)
            
            # Stil ekle
            header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            # Özet başlık satırını stilize et (satır 3 - DataFrame'in header'ı)
            for col in range(1, 3):  # A, B sütunları
                cell = worksheet.cell(row=3, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Hisse bazında başlık satırını stilize et
            if not df_profits.empty:
                header_row = start_row + 1
                for col in range(1, 5):  # A, B, C, D sütunları
                    cell = worksheet.cell(row=header_row, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # openpyxl yoksa basit versiyon
            summary_df.to_excel(writer, sheet_name='Kar/Zarar Analizi', index=False)
            if not df_profits.empty:
                df_profits.to_excel(writer, sheet_name='Kar/Zarar Analizi', index=False, startrow=len(summary_df) + 3)
        
        # 3. Özet sayfası (geriye dönük uyumluluk için)
        summary_data = {
            'Metrik': ['Toplam İşlem Sayısı', 'Toplam Alış', 'Toplam Satış', 'Toplam Kar/Zarar (TL)'],
            'Değer': [
                total_transactions,
                f"{total_buy:,.2f} TL",
                f"{total_sell:,.2f} TL",
                f"{total_profit_loss:+,.2f} TL"
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Özet', index=False)
    
    output.seek(0)
    return output.getvalue()

def export_transactions_to_csv(transactions):
    """İşlemleri CSV formatında export et"""
    if transactions:
        df_transactions = pd.DataFrame(transactions)
        if not df_transactions.empty and 'date' in df_transactions.columns:
            df_transactions = df_transactions[['date', 'symbol', 'type', 'quantity', 'price', 'total_value']]
            df_transactions['date'] = pd.to_datetime(df_transactions['date']).dt.strftime('%d.%m.%Y')
            df_transactions.columns = ['Tarih', 'Hisse', 'İşlem Tipi', 'Adet', 'Birim Fiyat (TL)', 'Toplam Tutar (TL)']
        else:
            df_transactions = pd.DataFrame(columns=['Tarih', 'Hisse', 'İşlem Tipi', 'Adet', 'Birim Fiyat (TL)', 'Toplam Tutar (TL)'])
    else:
        df_transactions = pd.DataFrame(columns=['Tarih', 'Hisse', 'İşlem Tipi', 'Adet', 'Birim Fiyat (TL)', 'Toplam Tutar (TL)'])
    
    # CSV formatında döndür
    return df_transactions.to_csv(index=False, encoding='utf-8-sig')

def analyze_bist100_trend(config, interval="1d"):
    """BIST 100 endeks trend analizi yapar"""
    try:
        loader = DataLoader(config)
        
        # Haftalık veri çek (son 3 ay)
        index_data = loader.get_index_data(period="3mo", interval="1wk")
        
        if index_data.empty:
            # Haftalık yoksa günlük veri ile haftalık hesapla
            index_data_daily = loader.get_index_data(period="3mo", interval="1d")
            if not index_data_daily.empty:
                # Günlük veriyi haftalık olarak grupla
                index_data = index_data_daily.resample('W').agg({
                    'open': 'first',
                    'high': 'max',
                    'low': 'min',
                    'close': 'last',
                    'volume': 'sum'
                })
        
        if index_data.empty or len(index_data) < 4:
            return None
        
        # Son 4 haftalık performans
        current_price = index_data['close'].iloc[-1]
        week_1_ago = index_data['close'].iloc[-2] if len(index_data) >= 2 else current_price
        week_2_ago = index_data['close'].iloc[-3] if len(index_data) >= 3 else current_price
        week_4_ago = index_data['close'].iloc[-4] if len(index_data) >= 4 else current_price
        
        # Haftalık getiriler
        return_1w = ((current_price / week_1_ago) - 1) * 100 if week_1_ago > 0 else 0
        return_2w = ((current_price / week_2_ago) - 1) * 100 if week_2_ago > 0 else 0
        return_4w = ((current_price / week_4_ago) - 1) * 100 if week_4_ago > 0 else 0
        
        # Trend belirleme
        # Eğer son 2 hafta düşüş varsa ve toplam düşüş %3'ten fazlaysa "DÜŞÜŞ"
        # Eğer son 2 hafta yükseliş varsa ve toplam yükseliş %3'ten fazlaysa "YÜKSELİŞ"
        # Diğer durumlarda "NÖTR"
        
        trend = "NÖTR"
        trend_strength = 0
        recommendation = None
        
        if return_2w < -3:  # Son 2 haftada %3'ten fazla düşüş
            trend = "DÜŞÜŞ"
            trend_strength = abs(return_2w)
            recommendation = "DÜŞÜKTEN_GİRİŞ"
        elif return_2w > 3:  # Son 2 haftada %3'ten fazla yükseliş
            trend = "YÜKSELİŞ"
            trend_strength = return_2w
            recommendation = "KAR_REALİZE"
        elif return_1w < -2:  # Son hafta %2'den fazla düşüş
            trend = "DÜŞÜŞ"
            trend_strength = abs(return_1w)
            recommendation = "DÜŞÜKTEN_GİRİŞ"
        elif return_1w > 2:  # Son hafta %2'den fazla yükseliş
            trend = "YÜKSELİŞ"
            trend_strength = return_1w
            recommendation = "KAR_REALİZE"
        
        return {
            'trend': trend,
            'trend_strength': trend_strength,
            'recommendation': recommendation,
            'current_price': current_price,
            'return_1w': return_1w,
            'return_2w': return_2w,
            'return_4w': return_4w,
            'week_1_ago': week_1_ago,
            'week_2_ago': week_2_ago,
            'week_4_ago': week_4_ago
        }
    except Exception as e:
        return None

def calculate_daily_recommendations(portfolio, config, interval="1d", investment_horizon="MEDIUM_TERM"):
    """Günlük önerileri hesapla"""
    recommendations = []
    cash = portfolio['cash']
    stocks = portfolio['stocks']
    
    # BIST 100 endeks trend analizi
    index_trend = analyze_bist100_trend(config, interval)
    
    # Önce mevcut pozisyonlar için analiz yap - SAT/ARTIR/TUT
    analyzed_positions = {}
    
    # İlk aşama: SAT önerilerini belirle (satıştan gelen parayı hesaplamak için)
    sell_recommendations = []
    for symbol in stocks:
        try:
            stock_info = stocks[symbol]
            quantity = stock_info['quantity']
            avg_cost = stock_info['avg_cost']
            
            # Model kontrolü ve otomatik eğitim
            symbol_name = symbol.replace('.IS', '')
            model_exists = False
            
            if os.path.exists('src/models'):
                model_files = [f for f in os.listdir('src/models') if f.endswith('.joblib')]
                model_exists = any(symbol_name in f for f in model_files)
            
            # Model yoksa otomatik eğit (sessiz mod)
            if not model_exists:
                try:
                    success, message = train_model_for_symbol(
                        symbol, config, 
                        progress_callback=None,
                        interval=interval, 
                        investment_horizon=investment_horizon
                    )
                except Exception as e:
                    pass
            
            # Hisse analizi yap
            result = analyze_single_stock(symbol, config, period="1y", interval=interval)
            
            if result is None:
                continue
            
            current_price = result['current_price']
            prediction = result.get('prediction')
            confidence = result.get('confidence', 0.5)
            
            # Mevcut değer
            current_value = quantity * current_price
            total_cost = quantity * avg_cost
            profit_loss = current_value - total_cost
            profit_loss_pct = (current_value / total_cost - 1) * 100 if total_cost > 0 else 0
            
            # Sadece SAT önerilerini kontrol et
            action = "TUT"
            if prediction == 0 and confidence > 0.60:
                if profit_loss_pct > 3:
                    action = "KISMEN SAT"
                elif profit_loss_pct < -7:
                    action = "SAT"
                elif profit_loss_pct < 0 and profit_loss_pct > -5:
                    action = "KISMEN SAT"
            elif prediction == 0 and 0.55 < confidence <= 0.60:
                if profit_loss_pct > 2:
                    action = "KISMEN SAT"
                elif profit_loss_pct < -5:
                    action = "KISMEN SAT"
            
            # SAT önerisi varsa kaydet
            if action in ["SAT", "KISMEN SAT"]:
                if action == "KISMEN SAT":
                    recommended_quantity = int(quantity / 2)
                    if recommended_quantity == 0:
                        recommended_quantity = quantity
                else:
                    recommended_quantity = quantity
                
                recommended_value = recommended_quantity * current_price
                
                sell_recommendations.append({
                    'symbol': symbol,
                    'action': action,
                    'recommended_value': recommended_value
                })
        except:
            continue
    
    # Satıştan gelecek toplam parayı hesapla
    total_sell_cash = sum([r['recommended_value'] for r in sell_recommendations])
    available_cash = cash + total_sell_cash  # Nakit + satıştan gelen para
    
    # İkinci aşama: Tüm pozisyonlar için öneri üret (ARTIR önerileri için available_cash kullan)
    for symbol in stocks:
        try:
            stock_info = stocks[symbol]
            quantity = stock_info['quantity']
            avg_cost = stock_info['avg_cost']
            
            # Model kontrolü ve otomatik eğitim
            symbol_name = symbol.replace('.IS', '')
            model_exists = False
            
            if os.path.exists('src/models'):
                model_files = [f for f in os.listdir('src/models') if f.endswith('.joblib')]
                model_exists = any(symbol_name in f for f in model_files)
            
            # Model yoksa otomatik eğit (sessiz mod)
            if not model_exists:
                try:
                    # Sessiz modda model eğit (mesaj gösterme)
                    success, message = train_model_for_symbol(
                        symbol, config, 
                        progress_callback=None,  # Sessiz mod
                        interval=interval, 
                        investment_horizon=investment_horizon
                    )
                except Exception as e:
                    # Model eğitimi başarısız olsa bile devam et
                    pass
            
            # Hisse analizi yap
            result = analyze_single_stock(symbol, config, period="1y", interval=interval)
            
            if result is None:
                continue
            
            current_price = result['current_price']
            prediction = result.get('prediction')
            confidence = result.get('confidence', 0.5)
            
            # Mevcut değer
            current_value = quantity * current_price
            total_cost = quantity * avg_cost
            profit_loss = current_value - total_cost
            profit_loss_pct = (current_value / total_cost - 1) * 100 if total_cost > 0 else 0
            
            # Öneri mantığı - Daha agresif ve öneri odaklı
            action = "TUT"
            action_reason = "Sinyal net değil - Bekle"
            
            # Teknik analiz sinyalleri de kontrol et (model yoksa veya güven düşükse)
            rsi = result.get('rsi', 50)
            trend_strength = result.get('trend_strength', '')
            volume_ratio = result.get('volume_ratio', 1.0)
            
            # Teknik analiz tabanlı sinyal hesapla
            technical_signal = None
            technical_confidence = 0.5
            
            if rsi < 35 and trend_strength == "Yükseliş" and volume_ratio > 1.2:
                technical_signal = 1  # AL
                technical_confidence = 0.60
            elif rsi > 65 and trend_strength == "Düşüş" and volume_ratio > 1.2:
                technical_signal = 0  # SAT
                technical_confidence = 0.60
            elif rsi < 30:
                technical_signal = 1  # AL (aşırı satım)
                technical_confidence = 0.55
            elif rsi > 70:
                technical_signal = 0  # SAT (aşırı alım)
                technical_confidence = 0.55
            
            # Model tahmini yoksa teknik analizi kullan
            if prediction is None and technical_signal is not None:
                prediction = technical_signal
                confidence = technical_confidence
            
            # Güçlü AL sinyali (>60% güven) - Eşik düşürüldü
            if prediction == 1 and confidence > 0.60:
                if profit_loss_pct > -5:  # %5'ten fazla zararda değilse
                    action = "ARTIR"
                    action_reason = f"🟢 Güçlü yükseliş sinyali - Fırsat (%{confidence*100:.0f} güven)"
                else:
                    action = "TUT"
                    action_reason = f"⚠️ Zararda pozisyon - Bekle (%{profit_loss_pct:.1f}%)"
            # Güçlü SAT sinyali (>60% güven)
            elif prediction == 0 and confidence > 0.60:
                if profit_loss_pct > 3:  # %3'ten fazla karda ise - eşik düşürüldü
                    action = "KISMEN SAT"
                    action_reason = f"🔴 Karı realize et - Düşüş sinyali (%{confidence*100:.0f} güven, %{profit_loss_pct:.1f} kar)"
                elif profit_loss_pct < -7:  # %7'den fazla zarardaysa - eşik düşürüldü
                    action = "SAT"
                    action_reason = f"⚠️ Stop Loss - Güçlü düşüş sinyali (%{confidence*100:.0f} güven, %{profit_loss_pct:.1f} zarar)"
                elif profit_loss_pct < 0 and profit_loss_pct > -5:
                    action = "KISMEN SAT"  # Küçük zararda kısmi satış öner
                    action_reason = f"💰 Küçük zarar - Kısmi stop loss (%{profit_loss_pct:.1f}%)"
                else:
                    action = "TUT"
                    action_reason = "📊 Pozisyon durumu normal"
            # Orta güvenli sinyal (55-60%) - Artık öneri veriyor
            elif prediction == 1 and 0.55 < confidence <= 0.60:
                if profit_loss_pct > -3:
                    action = "ARTIR"
                    action_reason = f"📈 Orta güvenli yükseliş sinyali - İhtiyatlı artırım (%{confidence*100:.0f} güven)"
                else:
                    action = "TUT"
                    action_reason = f"📈 Yükseliş ama zararda - Bekle (%{confidence*100:.0f} güven)"
            elif prediction == 0 and 0.55 < confidence <= 0.60:
                if profit_loss_pct > 2:  # Biraz karda ise
                    action = "KISMEN SAT"
                    action_reason = f"📉 Orta güvenli düşüş riski - İhtiyatlı satış (%{confidence*100:.0f} güven)"
                elif profit_loss_pct < -5:
                    action = "KISMEN SAT"
                    action_reason = f"📉 Küçük stop loss önerisi (%{profit_loss_pct:.1f}%)"
                else:
                    action = "TUT"
                    action_reason = f"📉 Düşüş riski orta seviyede - İzle (%{confidence*100:.0f} güven)"
            # Hafif sinyal (50-55%) - En azından bilgi ver
            elif prediction == 1 and 0.50 < confidence <= 0.55:
                if profit_loss_pct > 0:
                    action = "TUT"
                    action_reason = f"📊 Hafif yükseliş eğilimi - Karda olduğun için bekle (%{confidence*100:.0f} güven)"
                else:
                    action = "TUT"
                    action_reason = f"📊 Hafif yükseliş ama zararda - Dikkatli takip et (%{confidence*100:.0f} güven)"
            elif prediction == 0 and 0.50 < confidence <= 0.55:
                if profit_loss_pct > 5:
                    action = "TUT"
                    action_reason = f"📊 Hafif düşüş riski ama iyi karda - Dikkatli takip et (%{confidence*100:.0f} güven)"
                elif profit_loss_pct < -3:
                    action = "TUT"
                    action_reason = f"📊 Hafif düşüş riski ve zararda - Dikkatli izle (%{confidence*100:.0f} güven)"
                else:
                    action = "TUT"
                    action_reason = f"📊 Hafif sinyal - Net değil (%{confidence*100:.0f} güven)"
            # Düşük güven veya sinyal yok - Teknik analizle öner
            else:
                if technical_signal == 1:
                    action = "TUT"
                    action_reason = f"📊 Teknik analiz: Aşırı satım bölgesinde - Dikkatli takip (RSI: {rsi:.1f})"
                elif technical_signal == 0:
                    if profit_loss_pct > 2:
                        action = "KISMEN SAT"
                        action_reason = f"📊 Teknik analiz: Aşırı alım bölgesinde - İhtiyatlı satış (RSI: {rsi:.1f})"
                    else:
                        action = "TUT"
                        action_reason = f"📊 Teknik analiz: Aşırı alım ama karda değilsin (RSI: {rsi:.1f})"
                else:
                    action = "TUT"
                    action_reason = "⏳ Sinyal belirsiz - Pozisyon koru"
            
            # Hesaplanacak miktar
            recommended_quantity = 0
            recommended_value = 0
            recommended_price = current_price
            
            if action == "ARTIR":
                # Nakit varsa veya satıştan gelen para varsa artırım öner
                if available_cash > 0:
                    # Maksimum %15 sermaye ile artırım - daha agresif
                    max_addition = available_cash * 0.15
                    recommended_value = min(max_addition, current_value * 0.4)  # Mevcut pozisyonun %40'ına kadar
                    recommended_quantity = int(recommended_value / current_price)
                    # En az 1 lot öner
                    if recommended_quantity == 0:
                        recommended_quantity = 1
                        recommended_value = recommended_quantity * current_price
                else:
                    # Nakit yok ve satıştan gelen para da yoksa ARTIR önerisini ekleme, TUT olarak devam et
                    action = "TUT"
                    recommended_quantity = 0
                    recommended_value = 0
                    action_reason = "Nakit olmadığı için pozisyon korunuyor"
            elif action == "KISMEN SAT":
                # Yarısını sat
                recommended_quantity = int(quantity / 2)
                if recommended_quantity == 0:
                    recommended_quantity = quantity  # Eğer 1 adetteyse tümünü sat
                recommended_value = recommended_quantity * current_price
            elif action == "SAT":
                # Tümünü sat
                recommended_quantity = quantity
                recommended_value = current_value
            
            # Hedef fiyat hesapla (tüm öneriler için)
            target_price = current_price
            target_days = 30
            target_min_date = ''
            target_max_date = ''
            
            try:
                volatility = result.get('volatility', 0.3)
                data = load_stock_data(symbol, period="1y", interval=interval, silent=True)
                
                if not data.empty and prediction is not None:
                    price_predictor = PriceTargetPredictor(config)
                    price_targets = price_predictor.calculate_price_targets(
                        current_price, 
                        prediction, 
                        confidence, 
                        volatility / 100 if volatility > 1 else volatility, 
                        data,
                        {}
                    )
                    
                    # Hedef fiyat bilgilerini ekle
                    target_price = price_targets['targets']['moderate']
                    time_targets = price_targets.get('time_targets', {})
                    moderate_time = time_targets.get('moderate', {})
                    
                    target_days = moderate_time.get('estimated_days', 30)
                    target_min_date = moderate_time.get('min_date', '')
                    target_max_date = moderate_time.get('max_date', '')
            except Exception as e:
                # Hata durumunda varsayılan değerler
                pass
            
            recommendations.append({
                'symbol': symbol,
                'current_price': current_price,
                'recommended_price': recommended_price,
                'target_price': target_price,
                'target_days': target_days,
                'target_min_date': target_min_date,
                'target_max_date': target_max_date,
                'quantity': quantity,
                'avg_cost': avg_cost,
                'current_value': current_value,
                'total_cost': total_cost,
                'profit_loss': profit_loss,
                'profit_loss_pct': profit_loss_pct,
                'prediction': prediction,
                'confidence': confidence,
                'action': action,
                'action_reason': action_reason,
                'recommended_quantity': recommended_quantity,
                'recommended_value': recommended_value,
                'signal_strength': result.get('score', 0),
                'result': result  # Analiz sonucunu ekle
            })
            
        except Exception as e:
            st.warning(f"❌ {symbol} analizi hatası: {str(e)}")
            continue
    
    # Yeni alım önerileri için satışlardan gelecek nakit'i hesapla
    remaining_cash = cash
    
    # İlk olarak önerilen işlemleri kontrol et
    for rec in recommendations:
        if rec['action'] == "ARTIR":
            remaining_cash -= rec['recommended_value']
        elif rec['action'] in ["SAT", "KISMEN SAT"]:
            remaining_cash += rec['recommended_value']  # Satıştan gelen para
    
    # Mevcut portföyde olmayan hisseler için öneriler
    all_stocks = get_all_bist_stocks()
    # 50K yerine daha esnek: en az satıştan gelen nakit varsa öner
    min_cash_for_new_stocks = max(30000, remaining_cash * 0.3)  # En az 30K veya mevcut nakitin %30'u
    
    # HISSE AVCISI TARZI - Tüm hisseleri analiz et ve en iyilerini seç
    if remaining_cash > min_cash_for_new_stocks:
        scored_candidates = []
        
        for symbol in all_stocks:
            if symbol not in stocks:  # Sadece portföyde olmayan hisseler
                try:
                    # Model kontrolü ve otomatik eğitim
                    symbol_name = symbol.replace('.IS', '')
                    model_exists = False
                    
                    if os.path.exists('src/models'):
                        model_files = [f for f in os.listdir('src/models') if f.endswith('.joblib')]
                        model_exists = any(symbol_name in f for f in model_files)
                    
                    # Model yoksa otomatik eğit (sessiz mod)
                    if not model_exists:
                        try:
                            # Sessiz modda model eğit (mesaj gösterme)
                            success, message = train_model_for_symbol(
                                symbol, config, 
                                progress_callback=None,  # Sessiz mod
                                interval=interval, 
                                investment_horizon=investment_horizon
                            )
                        except Exception as e:
                            # Model eğitimi başarısız olsa bile devam et
                            pass
                    
                    result = analyze_single_stock(symbol, config, period="1y", interval=interval, silent=True)
                    
                    if result is None:
                        continue
                    
                    prediction = result.get('prediction')
                    confidence = result.get('confidence', 0.5)
                    score = result.get('score', 0)
                    current_price = result['current_price']
                    
                    # AL sinyali olan hisseleri düşün - güven eşiğini %48'e düşür (daha fazla seçenek için)
                    if prediction == 1 and confidence > 0.48:
                        scored_candidates.append({
                            'symbol': symbol,
                            'current_price': current_price,
                            'prediction': prediction,
                            'confidence': confidence,
                            'score': score,
                            'result': result
                        })
                
                except:
                    continue
        
        # En yüksek skorlu olanları seç - HISSE AVCISI MANTIGI
        scored_candidates.sort(key=lambda x: x['score'], reverse=True)
        
        # Satıştan gelen toplam parayı hesapla
        sell_cash = sum([r['recommended_value'] for r in recommendations if r['action'] in ['SAT', 'KISMEN SAT']])
        total_cash_after_sell = cash + sell_cash
        
        # RİSK YÖNETİMİ: En az 3-5 hisse seç (çeşitlendirme için zorunlu)
        # Sadece 1-2 hisse bulunursa bile, güven eşiğini düşürerek daha fazla hisse bul
        
        # Önce en iyi 7-10 hisseyi dene (daha fazla seçenek)
        num_candidates = min(10, len(scored_candidates))
        top_candidates = scored_candidates[:num_candidates]
        
        # RİSK YÖNETİMİ: En az 3-5 hisse bulunmalı - eğer yoksa daha fazla hisse bul
        if len(top_candidates) < 5:
            # Tüm scored_candidates listesini kontrol et (zaten %50 güven eşiği ile geldi)
            # Skorlarına göre sırala ve en iyi olanları ekle
            all_candidates = sorted(scored_candidates, key=lambda x: x['score'], reverse=True)
            
            # En az 5 hisse bulunana kadar ekle
            needed = 5 - len(top_candidates)
            additional_candidates = all_candidates[len(top_candidates):len(top_candidates) + needed + 3]  # Biraz fazla seç
            
            # Eğer hala yeterli değilse, güven eşiğini %45'e düşür ve yeniden ara
            if len(top_candidates) + len(additional_candidates) < 5:
                # Tüm hisseleri yeniden kontrol et (daha düşük eşikle)
                checked_symbols = set([c['symbol'] for c in top_candidates + additional_candidates])
                for symbol in all_stocks:
                    if symbol not in stocks and symbol not in checked_symbols:
                        try:
                            # Model kontrolü ve otomatik eğitim
                            symbol_name = symbol.replace('.IS', '')
                            model_exists = False
                            
                            if os.path.exists('src/models'):
                                model_files = [f for f in os.listdir('src/models') if f.endswith('.joblib')]
                                model_exists = any(symbol_name in f for f in model_files)
                            
                            # Model yoksa otomatik eğit (sessiz mod)
                            if not model_exists:
                                try:
                                    success, message = train_model_for_symbol(
                                        symbol, config, 
                                        progress_callback=None,
                                        interval=interval, 
                                        investment_horizon=investment_horizon
                                    )
                                except Exception as e:
                                    pass
                            
                            result = analyze_single_stock(symbol, config, period="1y", interval=interval, silent=True)
                            if result is None:
                                continue
                            
                            prediction = result.get('prediction')
                            confidence = result.get('confidence', 0.5)
                            if prediction == 1 and confidence > 0.45:  # %45'e düşür
                                additional_candidates.append({
                                    'symbol': symbol,
                                    'current_price': result['current_price'],
                                    'prediction': prediction,
                                    'confidence': confidence,
                                    'score': result.get('score', 0),
                                    'result': result
                                })
                                checked_symbols.add(symbol)
                                if len(top_candidates) + len(additional_candidates) >= 7:
                                    break
                        except:
                            continue
            
            top_candidates.extend(additional_candidates[:needed])
            
            # Son durumda skorlarına göre sırala ve en iyi 5-7 hisseyi seç
            top_candidates = sorted(top_candidates, key=lambda x: x['score'], reverse=True)[:min(7, len(top_candidates))]
            
            # Minimum 3 hisse garantisi - eğer hala azsa en iyi olanları zorla seç
            if len(top_candidates) < 3 and len(scored_candidates) >= 3:
                top_candidates = scored_candidates[:3]  # En az 3 hisse zorla seç
        
        for idx, candidate in enumerate(top_candidates):
            symbol = candidate['symbol']
            current_price = candidate['current_price']
            confidence = candidate['confidence']
            score = candidate['score']
            
            # Dinamik tahsis: RİSK YÖNETİMİ - Çeşitlendirme önemli!
            num_candidates = len(top_candidates)
            
            if num_candidates == 1:
                # Sadece 1 hisse varsa bile - maksimum %30 (risk yönetimi)
                allocation_pct = 0.30
            elif num_candidates == 2:
                # 2 hisse varsa - her birine %25-30
                allocation_pct = 0.30
            elif num_candidates == 3:
                # 3 hisse varsa - her birine %20 (60% toplam)
                allocation_pct = 0.20
            elif num_candidates <= 5:
                # 4-5 hisse varsa - her birine %15-18
                allocation_pct = 0.18
            else:
                # 6+ hisse varsa - her birine %12-15
                allocation_pct = 0.15
            
            # Sermayenin tahsis edilen yüzdesi kadar öner
            recommended_value = min(
                total_cash_after_sell * allocation_pct,  # Toplam sermayenin tahsis yüzdesi
                remaining_cash  # Kalan nakitten fazla olmamalı
            )
            
            # En az fiyatın 50 katı kadar öner
            min_value = current_price * 50
            recommended_value = max(min_value, recommended_value)
            
            # Eğer son hisseyse ve hala çok nakit kaldıysa, makul bir şekilde kullan
            if idx == len(top_candidates) - 1:
                # Son hisse için kalan paranın %50'sini kullan (maksimum)
                if remaining_cash > recommended_value * 1.5:  # Eğer kalan nakit önerilenden 1.5x fazlaysa
                    # Ama toplam tahsisi %30'u geçmesin (risk yönetimi)
                    max_allocation = total_cash_after_sell * 0.30
                    recommended_value = min(remaining_cash * 0.50, max_allocation)
            
            recommended_quantity = int(recommended_value / current_price)
            
            if recommended_quantity > 0 and remaining_cash >= recommended_value:
                # Hedef fiyat ve tarih bilgilerini hesapla
                try:
                    result_data = candidate['result']
                    volatility = result_data.get('volatility', 0.3)
                    data = load_stock_data(symbol, period="1y", interval=interval, silent=True)
                    
                    if not data.empty:
                        price_predictor = PriceTargetPredictor(config)
                        price_targets = price_predictor.calculate_price_targets(
                            current_price, 
                            candidate['prediction'], 
                            confidence, 
                            volatility / 100 if volatility > 1 else volatility, 
                            data,
                            {}
                        )
                        
                        # Hedef fiyat bilgilerini ekle
                        target_moderate = price_targets['targets']['moderate']
                        time_targets = price_targets.get('time_targets', {})
                        moderate_time = time_targets.get('moderate', {})
                        
                        estimated_days = moderate_time.get('estimated_days', 30)
                        min_date = moderate_time.get('min_date', '')
                        max_date = moderate_time.get('max_date', '')
                    else:
                        target_moderate = current_price * 1.10  # Varsayılan %10 artış
                        estimated_days = 30
                        min_date = ''
                        max_date = ''
                except Exception as e:
                    # Hata durumunda varsayılan değerler
                    target_moderate = current_price * 1.10
                    estimated_days = 30
                    min_date = ''
                    max_date = ''
                
                recommendations.append({
                    'symbol': symbol,
                    'current_price': current_price,
                    'recommended_price': current_price,
                    'target_price': target_moderate,
                    'target_days': estimated_days,
                    'target_min_date': min_date,
                    'target_max_date': max_date,
                    'quantity': 0,
                    'avg_cost': 0,
                    'current_value': 0,
                    'total_cost': 0,
                    'profit_loss': 0,
                    'profit_loss_pct': 0,
                    'prediction': candidate['prediction'],
                    'confidence': confidence,
                    'action': "YENİ AL",
                    'action_reason': f"Güçlü yükseliş fırsatı (%{confidence*100:.0f} güven)",
                    'recommended_quantity': recommended_quantity,
                    'recommended_value': recommended_value,
                    'signal_strength': score,
                    'from_sell': True,
                    'allocation_pct': allocation_pct,
                    'result': candidate.get('result', {})  # Analiz sonucunu ekle
                })
                remaining_cash -= recommended_value
                
                # Kalan nakdi azaldıysa dur
                if remaining_cash < total_cash_after_sell * 0.05:  # %5'in altındaysa dur
                    break
    
    return recommendations, index_trend

def show_portfolio_manager_tab(config, interval="1d", investment_horizon="MEDIUM_TERM"):
    """Robot Portföy Yöneticisi sekmesini göster"""
    
    # Authentication kontrolü
    db = get_db()
    user_id = require_auth(db)
    
    if not user_id:
        st.stop()
    
    # Başlık
    st.markdown("""
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                padding: 20px; 
                border-radius: 15px; 
                margin-bottom: 30px;
                text-align: center;">
        <h1 style="color: white; margin: 0;">🤖 Robot Portföy Yöneticisi</h1>
        <p style="color: white; margin: 10px 0 0 0; font-size: 1.1em;">
            AI Destekli Günlük Yatırım Kararları
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Bilgilendirme
    with st.expander("ℹ️ Robot Portföy Yöneticisi Nedir?", expanded=False):
        st.markdown("""
        **Robot Portföy Yöneticisi**, gün sonunda hisse senedi borsasını kapattıktan sonra yapmanız gereken işlemleri gösterir.
        
        #### 🎯 Nasıl Çalışır?
        - **Nakit paranızı** girin (örn: 100,000 TL)
        - **Mevcut portföyünüzü** ekleyin (hangi hisselerden kaç adet var, ortalama maliyetiniz ne?)
        - Sistem **AI analizi** yaparak günlük öneriler sunar
        
        #### 🤖 Öneri Mantığı:
        - **AL/ARTIR:** Güçlü yükseliş sinyali varsa - Fırsat kaçmasın
        - **SAT:** Karı realize et veya zararı durdur - Risk yönetimi
        - **TUT:** Sinyal net değil - Zırt pırt değişiklik yapma
        
        #### 💡 Özellikleri:
        - Gereksiz işlem önleme (zırt pırt değişiklik yapmaz)
        - Maliyet bazlı karar (aldığınız fiyata göre)
        - Net öneriler (ne kadar al, ne kadar sat - çok net)
        """)
    
    # Portföy yükle
    portfolio = load_portfolio(user_id)
    
    # === MIGRATION BİLGİSİ ===
    # Eski JSON transaction'ları kontrol et ve migrate et
    migration_key = f'migration_done_{user_id}'
    migration_file = os.path.join(TRANSACTIONS_DIR, f'.migration_done_{user_id}.json')
    
    # Migration durumunu kontrol et (dosyadan veya session state'ten)
    migration_done = False
    if os.path.exists(migration_file):
        migration_done = True
    elif st.session_state.get(migration_key, False):
        migration_done = True
    
    # Eski transaction dosyalarını kontrol et
    old_transaction_files = []
    if os.path.exists(TRANSACTIONS_DIR):
        old_transaction_files = [f for f in os.listdir(TRANSACTIONS_DIR) 
                               if f.startswith("transactions_") and f.endswith(".json")]
    
    # Mevcut transaction'ları yükle
    transactions = load_user_transactions(user_id)
    
    # Migration yapılmadıysa ve eski dosyalar varsa göster
    if old_transaction_files and not migration_done and len(transactions) == 0:
        st.info("""
        **📦 Eski Transaction Verileri Bulundu**
        
        Eski sistemden transaction verileriniz bulundu. Bu verileri yeni veritabanı sistemine aktarmak ister misiniz?
        
        **Not:** Migration işlemi sadece bir kez çalıştırılmalıdır.
        """)
        
        if st.button("🔄 Eski Transaction'ları Aktar", type="primary", key="migrate_button"):
            with st.spinner("Transaction'lar aktarılıyor..."):
                db = get_db()
                total_migrated = 0
                migration_details = []
                
                for json_file in old_transaction_files:
                    old_user_id = json_file.replace("transactions_", "").replace(".json", "")
                    json_path = os.path.join(TRANSACTIONS_DIR, json_file)
                    
                    # Dosya var mı kontrol et
                    if not os.path.exists(json_path):
                        st.warning(f"⚠️ Dosya bulunamadı: {json_file}")
                        continue
                    
                    # Dosya içeriğini kontrol et
                    try:
                        with open(json_path, 'r', encoding='utf-8') as f:
                            test_data = json.load(f)
                            file_count = len(test_data) if isinstance(test_data, list) else 0
                    except Exception as e:
                        st.error(f"❌ Dosya okunamadı: {json_file} - {str(e)}")
                        continue
                    
                    # Migration yap (duplicate kontrolü otomatik yapılır)
                    migrated_count = db.migrate_json_transactions(old_user_id, user_id, json_path)
                    total_migrated += migrated_count
                    migration_details.append(f"{json_file}: {migrated_count}/{file_count} transaction (duplicate'ler atlandı)")
                
                if total_migrated > 0:
                    # Migration flag'ini kalıcı olarak kaydet
                    st.session_state[migration_key] = True
                    os.makedirs(TRANSACTIONS_DIR, exist_ok=True)
                    with open(migration_file, 'w', encoding='utf-8') as f:
                        json.dump({'migrated': True, 'count': total_migrated, 'user_id': user_id, 'details': migration_details}, f, indent=2)
                    
                    # Transaction'ları tekrar yükle ve kontrol et
                    transactions_after = load_user_transactions(user_id)
                    
                    st.success(f"✅ {total_migrated} transaction başarıyla aktarıldı!")
                    if len(transactions_after) > 0:
                        st.info(f"📊 Veritabanında {len(transactions_after)} transaction bulundu.")
                    else:
                        st.warning("⚠️ Transaction'lar aktarıldı ancak yüklenemedi. Sayfayı yenileyin.")
                    
                    st.rerun()
                else:
                    st.error("❌ Hiç transaction aktarılamadı. Lütfen dosya formatını kontrol edin.")
                    if old_transaction_files:
                        st.code(f"Dosyalar: {', '.join(old_transaction_files)}")
    
    # Transaction'ları tekrar yükle (migration'dan sonra güncel olması için)
    transactions = load_user_transactions(user_id)
    
    # === PORTFÖY GİRİŞİ ===
    st.markdown("---")
    st.markdown("### 📝 Portföy Bilgileri")
    
    col1, col2 = st.columns(2)
    
    with col1:
        cash = st.number_input(
            "💵 Nakit Para (TL):",
            min_value=0.0,
            value=float(portfolio.get('cash', 0)),
            step=1000.0,
            format="%.2f",
            help="Borsaya yatırılacak nakit parayı girin"
        )
    
    with col2:
        st.metric(
            "📊 Mevcut Portföy",
            f"{len(portfolio['stocks'])} hisse",
            help="Portföydeki hisse sayısı"
        )
    
    # === HİSSE EKLEME/SİLME ===
    st.markdown("#### 📋 Portföydeki Hisseler")
    
    # İşlemlerden pozisyon hesaplamak için transaction'ları yükle
    user_transactions = load_user_transactions(user_id)
    
    # Alış-satış işlemlerinden portföyü güncelle butonu
    if user_transactions:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.info("💡 Alış-satış işlemlerinizden kalan pozisyonları otomatik olarak portföye ekleyebilirsiniz.")
        with col2:
            if st.button("🔄 İşlemlerden Portföyü Güncelle", type="primary", use_container_width=True):
                # Alış-satış işlemlerinden kalan pozisyonları hesapla
                calculated_positions = calculate_remaining_positions(user_transactions)
                
                if calculated_positions:
                    # Portföyü güncelle (mevcut pozisyonları koru, yeni olanları ekle veya güncelle)
                    updated_count = 0
                    for symbol, position_data in calculated_positions.items():
                        portfolio['stocks'][symbol] = {
                            'quantity': position_data['quantity'],
                            'avg_cost': position_data['avg_cost']
                        }
                        updated_count += 1
                    
                    save_portfolio(portfolio, user_id)
                    st.success(f"✅ {updated_count} hisse pozisyonu alış-satış işlemlerinden hesaplanarak portföye eklendi/güncellendi!")
                    st.rerun()
                else:
                    st.warning("⚠️ Alış-satış işlemlerinden kalan pozisyon bulunamadı.")
    
    all_stocks = get_all_bist_stocks()
    
    # Yeni hisse ekleme
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        new_stock = st.selectbox(
            "Hisse Seç:",
            all_stocks,
            key="new_stock_select"
        )
    
    with col2:
        stock_quantity = st.number_input(
            "Adet:",
            min_value=1,
            value=100,
            step=10,
            key="stock_quantity"
        )
    
    with col3:
        stock_cost = st.number_input(
            "Ortalama Maliyet (TL):",
            min_value=0.01,
            value=100.0,
            step=0.10,
            key="stock_cost"
        )
    
    if st.button("➕ Hisse Ekle", type="primary"):
        portfolio['stocks'][new_stock] = {
            'quantity': int(stock_quantity),
            'avg_cost': float(stock_cost)
        }
        save_portfolio(portfolio, user_id)
        st.success(f"✅ {new_stock} eklendi!")
        st.rerun()
    
    # Mevcut hisseleri göster
    if portfolio['stocks']:
        st.markdown("**Mevcut Pozisyonlar:**")
        for symbol, info in portfolio['stocks'].items():
            col1, col2, col3, col4, col5 = st.columns([2, 1, 1, 1, 1])
            
            with col1:
                st.write(f"📊 {symbol}")
            
            with col2:
                st.write(f"{info['quantity']:.0f} adet")
            
            with col3:
                st.write(f"{info['avg_cost']:.2f} TL")
            
            with col4:
                total_cost = info['quantity'] * info['avg_cost']
                st.write(f"{format_currency(total_cost):,.2f} TL")
            
            with col5:
                if st.button(f"❌", key=f"del_{symbol}", help="Sil"):
                    del portfolio['stocks'][symbol]
                    save_portfolio(portfolio, user_id)
                    st.rerun()
    
    # Nakit güncelle
    portfolio['cash'] = cash
    save_portfolio(portfolio, user_id)
    
    # === GÜNLÜK ÖNERİLER ===
    if portfolio['cash'] > 0 or portfolio['stocks']:
        st.markdown("---")
        
        # Kompakt başlık
        st.markdown("### 🤖 Günlük Portföy Analizi")
        st.info("💡 AI senin için bugün ne yapman gerektiğine karar verdi!")
        
        # Portföy özeti
        total_portfolio_value = sum([s['quantity'] * s['avg_cost'] for s in portfolio['stocks'].values()])
        total_cash = portfolio['cash']
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("💵 Toplam Nakit", f"{format_currency(total_cash):,.2f} TL")
        with col2:
            st.metric("📊 Portföy Değeri", f"{format_currency(total_portfolio_value):,.2f} TL")
        with col3:
            st.metric("💰 Toplam Sermaye", f"{format_currency(total_cash + total_portfolio_value):,.2f} TL")
        
        st.markdown("---")
        
        # Büyük analiz başlat butonu ve durumu
        col1, col2, col3 = st.columns([1, 3, 1])
        with col2:
            # Loading durumunu kontrol et
            is_analyzing = st.session_state.get('analyze_clicked', False)
            
            if is_analyzing:
                # Analiz yapılırken butonu göster (devre dışı görünüm)
                st.button("⏳ ANALİZ YAPILIYOR...", disabled=True, use_container_width=True)
            else:
                analyze_button = st.button("🤖 GÜNLÜK ANALİZİ BAŞLAT - AI ÖNERİLERİNİ AL", type="primary", use_container_width=True)
                if analyze_button:
                    st.session_state.analyze_clicked = True
                    st.rerun()  # Sayfayı yenileyerek loading göster
        
        # Progress mesajı - Sadece butona basıldığında analiz yap
        if st.session_state.get('analyze_clicked', False):
            # UI Friendly loading
            with st.spinner("🔍 AI analizi başladı..."):
                progress_bar = st.progress(0)
                progress_status = st.empty()
                
                # Simüle edilmiş progress
                progress_bar.progress(10)
                progress_status.text("📊 Mevcut pozisyonlar analiz ediliyor...")
                
                progress_bar.progress(20)
                progress_status.text("🤖 Model durumu kontrol ediliyor...")
                
                progress_bar.progress(40)
                progress_status.text("💰 Satış önerileri hesaplanıyor...")
                
                progress_bar.progress(60)
                progress_status.text("🔍 Yeni hisse fırsatları taraniyor...")
                
                progress_bar.progress(80)
                progress_status.text("🎯 En iyi fırsatlar seçiliyor...")
                
                result = calculate_daily_recommendations(
                    portfolio, config, interval, investment_horizon
                )
                
                progress_bar.progress(90)
                progress_status.text("📋 Öneriler hazırlanıyor...")
                
                # Sonuçları sakla
                if isinstance(result, tuple):
                    recommendations, index_trend = result
                else:
                    recommendations = result
                    index_trend = None
                st.session_state.last_recommendations = recommendations
                st.session_state.last_index_trend = index_trend
                st.session_state.analyze_clicked = False  # Analiz tamamlandı
                
                progress_bar.progress(100)
                progress_status.text("✅ Analiz tamamlandı!")
                
        else:
            # Önceki sonuçları kullan veya boş
            recommendations = st.session_state.get('last_recommendations', [])
            index_trend = st.session_state.get('last_index_trend', None)
            
            # Eğer hiç analiz yoksa kullanıcıya bilgi ver
            if not recommendations:
                st.warning("ℹ️ Analizi başlatmak için yukarıdaki butona basın.")
        
        if recommendations:
            # Özet kartları
            actions = [r['action'] for r in recommendations]
            st.metrics = {"Önerilen İşlem": f"{len(recommendations)} hisse"}
            
            # BIST 100 Endeks Trend Analizi - ÖNEMLİ BİLGİ
            if index_trend:
                st.markdown("---")
                st.markdown("#### 📊 BIST 100 Endeks Trend Analizi")
                
                trend = index_trend['trend']
                trend_strength = index_trend['trend_strength']
                return_1w = index_trend['return_1w']
                return_2w = index_trend['return_2w']
                return_4w = index_trend['return_4w']
                current_index_price = index_trend['current_price']
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    if trend == "DÜŞÜŞ":
                        st.error(f"📉 **{trend}**")
                        st.metric("Trend Gücü", f"%{trend_strength:.2f}")
                    elif trend == "YÜKSELİŞ":
                        st.success(f"📈 **{trend}**")
                        st.metric("Trend Gücü", f"%{trend_strength:.2f}")
                    else:
                        st.info(f"➡️ **{trend}**")
                        st.metric("Trend Gücü", f"%{trend_strength:.2f}")
                
                with col2:
                    st.metric("Son 1 Hafta", f"%{return_1w:+.2f}")
                
                with col3:
                    st.metric("Son 2 Hafta", f"%{return_2w:+.2f}")
                
                with col4:
                    st.metric("Son 4 Hafta", f"%{return_4w:+.2f}")
                
                # Endeks önerisi
                if index_trend['recommendation'] == "DÜŞÜKTEN_GİRİŞ":
                    st.warning(f"""
                    **💡 Endeks Önerisi: Düşükten Giriş Fırsatı**
                    
                    BIST 100 endeksi son {2 if return_2w < -3 else 1} haftada **%{abs(return_2w if return_2w < -3 else return_1w):.2f}** düştü. 
                    Bu, yedek paranızla düşükten giriş yapmak için bir fırsat olabilir.
                    
                    **Öneri:** Yedek paranızın bir kısmını (%20-30) kullanarak güçlü hisselere düşükten giriş yapmayı düşünün.
                    """)
                elif index_trend['recommendation'] == "KAR_REALİZE":
                    st.info(f"""
                    **💡 Endeks Önerisi: Kar Realizasyonu**
                    
                    BIST 100 endeksi son {2 if return_2w > 3 else 1} haftada **%{return_2w if return_2w > 3 else return_1w:.2f}** yükseldi. 
                    Bu, karlı pozisyonlarınızdan kar realize etmek için uygun bir zaman olabilir.
                    
                    **Öneri:** Karlı pozisyonlarınızın bir kısmından (%20-30) kar realize ederek nakit pozisyonunuzu güçlendirin.
                    """)
                else:
                    st.info(f"""
                    **💡 Endeks Durumu: Nötr**
                    
                    BIST 100 endeksi son haftalarda belirgin bir trend göstermiyor. 
                    Normal portföy yönetimi stratejilerinizi uygulayabilirsiniz.
                    """)
                
                st.markdown("---")
            
            # İşlem grupları - İŞLEM SIRASINA GÖRE (SAT → AL → ARTIR → TUT)
            sell_actions = [r for r in recommendations if r['action'] in ['SAT', 'KISMEN SAT']]
            new_buy_actions = [r for r in recommendations if r['action'] == 'YENİ AL']
            increase_actions = [r for r in recommendations if r['action'] == 'ARTIR']
            hold_actions = [r for r in recommendations if r['action'] == 'TUT']
            
            # Tüm buy actions (gösterim için)
            buy_actions = new_buy_actions + increase_actions
            
            # Satıştan gelen nakit bilgisi - BÜYÜK VE ÇARPICI
            total_sell_cash = sum([r['recommended_value'] for r in sell_actions])
            if total_sell_cash > 0:
                # Bu nakitle nereye yatırım yapılacağını göster
                new_buy_from_sell = [r for r in buy_actions if r.get('from_sell', False)]
                
                if new_buy_from_sell:
                    total_buy_from_sell = sum([r['recommended_value'] for r in new_buy_from_sell])
                    usage_pct = (total_buy_from_sell / total_sell_cash * 100) if total_sell_cash > 0 else 0
                    
                    # BÜYÜK BANNER - ÖNEMLİ BİLGİ
                    # Kompakt banner
                    st.success(f"💰 **SATIŞLARDAN GELECEK:** {format_currency(total_sell_cash):,.2f} TL\n"
                              f"💡 Bu para ile {format_currency(total_buy_from_sell):,.2f} TL ({usage_pct:.0f}%) tutarında **{len(new_buy_from_sell)} yeni hisse** önerisi hazırlandı!")
            
            # 1. SATIŞ ÖNERİLERİ (ÖNCE SAT - PARA ÇIKACAK)
            if sell_actions:
                st.markdown("#### 🔴 1️⃣ SATIŞ ÖNERİLERİ (Önce bunları yap)")
                for rec_idx, rec in enumerate(sell_actions):
                    col1, col2, col3 = st.columns([3, 2, 2])
                    
                    with col1:
                        st.markdown(f"### 📉 {rec['symbol']}")
                        if rec['action'] == 'SAT':
                            st.error("⚠️ Tüm Pozisyonu Sat")
                        else:
                            st.warning("⚠️ Kısmi Satış")
                    
                    with col2:
                        st.metric("💰 Güncel Fiyat", f"{format_currency(rec['current_price']):,.2f} TL")
                        st.metric("📦 Önerilen Satış", f"{rec['recommended_quantity']:.0f} adet")
                        st.metric("📊 Mevcut", f"{rec['quantity']:.0f} adet")
                        # Hedef fiyat bilgisi
                        target_price = rec.get('target_price', rec['current_price'])
                        if target_price != rec['current_price']:
                            potential_return = ((target_price - rec['current_price']) / rec['current_price']) * 100
                            st.metric("🎯 Hedef Fiyat", f"{format_currency(target_price):,.2f} TL", 
                                    delta=f"%{potential_return:+.1f}")
                    
                    with col3:
                        st.metric("💵 Satış Tutarı", f"{format_currency(rec['recommended_value']):,.2f} TL")
                        st.metric("📈 Kar/Zarar", f"{rec['profit_loss_pct']:+.2f}%", 
                                delta=f"{format_currency(rec['profit_loss']):+,.2f} TL")
                        st.caption(f"Ortalama maliyet: {format_currency(rec['avg_cost']):,.2f} TL")
                        # Hedef tarih bilgisi
                        target_days = rec.get('target_days', 30)
                        target_min_date = rec.get('target_min_date', '')
                        target_max_date = rec.get('target_max_date', '')
                        if target_min_date and target_max_date:
                            st.caption(f"📅 Hedef Tarih: {target_min_date} - {target_max_date}")
                        elif target_days and target_price != rec['current_price']:
                            target_date = (datetime.now() + timedelta(days=target_days)).strftime('%d.%m.%Y')
                            st.caption(f"📅 Tahmini Süre: ~{target_days} gün ({target_date})")
                    st.markdown(f"**💬 Ne Yapılacak:** {rec['recommended_quantity']:.0f} adet {rec['symbol']} sat, {format_currency(rec['recommended_value']):,.2f} TL al")
                    # Sadece neden varsa göster
                    if rec.get('action_reason'):
                        st.caption(f"💡 {rec['action_reason']}")
                    st.divider()
            
            # 2. YENİ ALIM ÖNERİLERİ (SATIŞTAN SONRA - YENİ HİSSELER)
            if new_buy_actions:
                st.markdown("#### 🟢 2️⃣ YENİ ALIM ÖNERİLERİ (Satıştan gelen parayla)")
                for rec in new_buy_actions:
                    is_from_sell = rec.get('from_sell', False)
                    col1, col2, col3 = st.columns([3, 2, 2])
                    
                    with col1:
                        st.markdown(f"### 📈 {rec['symbol']}")
                        if is_from_sell:
                            st.success("🆕 Yeni Alım (Satıştan gelen parayla)")
                        else:
                            st.success("🆕 Yeni Alım Önerisi")
                    
                    with col2:
                        st.metric("💰 Güncel Fiyat", f"{format_currency(rec['current_price']):,.2f} TL")
                        st.metric("📦 Önerilen Miktar", f"{rec['recommended_quantity']:.0f} adet")
                        # Hedef fiyat bilgisi
                        target_price = rec.get('target_price', rec['current_price'])
                        potential_return = ((target_price - rec['current_price']) / rec['current_price']) * 100
                        st.metric("🎯 Hedef Fiyat", f"{format_currency(target_price):,.2f} TL", 
                                delta=f"%{potential_return:+.1f} getiri")
                    
                    with col3:
                        st.metric("💵 İşlem Tutarı", f"{format_currency(rec['recommended_value']):,.2f} TL")
                        st.metric("🎯 Güven Skoru", f"%{rec['confidence']*100:.0f}")
                        allocation = rec.get('allocation_pct', 0)
                        if allocation > 0:
                            st.caption(f"📊 Portföy Tahsisi: %{allocation*100:.0f}")
                        
                        # Hedef tarih bilgisi
                        target_days = rec.get('target_days', 30)
                        target_min_date = rec.get('target_min_date', '')
                        target_max_date = rec.get('target_max_date', '')
                        if target_min_date and target_max_date:
                            st.caption(f"📅 Hedef Tarih: {target_min_date} - {target_max_date}")
                        elif target_days:
                            target_date = (datetime.now() + timedelta(days=target_days)).strftime('%d.%m.%Y')
                            st.caption(f"📅 Tahmini Süre: ~{target_days} gün ({target_date})")
                        
                        if is_from_sell:
                            st.markdown(f"**💬 Ne Yapılacak:** Satıştan gelen parayla {rec['recommended_quantity']:.0f} adet {rec['symbol']} al, {format_currency(rec['recommended_value']):,.2f} TL harca")
                        else:
                            st.markdown(f"**💬 Ne Yapılacak:** {rec['recommended_quantity']:.0f} adet {rec['symbol']} al, {format_currency(rec['recommended_value']):,.2f} TL harca")
                    
                    # Sadece neden varsa göster
                    if rec.get('action_reason'):
                        st.caption(f"💡 {rec['action_reason']}")
                    
                    # Neden AL dediğinin detaylı özeti
                    st.markdown("**📋 Neden AL Önerisi:**")
                    from dashboard_portfolio_export import generate_buy_reasons
                    buy_reasons = generate_buy_reasons(rec)
                    for reason in buy_reasons:
                        st.markdown(f"  • {reason}")
                    
                    st.divider()
            
            # 3. MEVCUT POZİSYON ARTIRIMI (OPTIONAL)
            # Sadece nakit varsa ve işlem tutarı > 0 olanları göster
            increase_actions_filtered = [r for r in increase_actions if r.get('recommended_value', 0) > 0]
            if increase_actions_filtered:
                st.markdown("#### 📈 3️⃣ MEVCUT POZİSYON ARTIRIMLARI")
                for rec in increase_actions_filtered:
                    col1, col2, col3 = st.columns([3, 2, 2])
                    
                    with col1:
                        st.markdown(f"### 📈 {rec['symbol']}")
                        st.success("📊 Mevcut Pozisyon Artırımı")
                    
                    with col2:
                        st.metric("💰 Güncel Fiyat", f"{format_currency(rec['current_price']):,.2f} TL")
                        st.metric("📦 Önerilen Miktar", f"{rec['recommended_quantity']:.0f} adet")
                        st.caption(f"Şu an: {rec['quantity']:.0f} adet")
                        # Hedef fiyat bilgisi
                        target_price = rec.get('target_price', rec['current_price'])
                        if target_price != rec['current_price']:
                            potential_return = ((target_price - rec['current_price']) / rec['current_price']) * 100
                            st.metric("🎯 Hedef Fiyat", f"{format_currency(target_price):,.2f} TL", 
                                    delta=f"%{potential_return:+.1f} getiri")
                    
                    with col3:
                        st.metric("💵 İşlem Tutarı", f"{format_currency(rec['recommended_value']):,.2f} TL")
                        st.metric("🎯 Güven Skoru", f"%{rec['confidence']*100:.0f}")
                        st.caption(f"Ortalama maliyet: {format_currency(rec['avg_cost']):,.2f} TL")
                        # Hedef tarih bilgisi
                        target_days = rec.get('target_days', 30)
                        target_min_date = rec.get('target_min_date', '')
                        target_max_date = rec.get('target_max_date', '')
                        if target_min_date and target_max_date:
                            st.caption(f"📅 Hedef Tarih: {target_min_date} - {target_max_date}")
                        elif target_days and target_price != rec['current_price']:
                            target_date = (datetime.now() + timedelta(days=target_days)).strftime('%d.%m.%Y')
                            st.caption(f"📅 Tahmini Süre: ~{target_days} gün ({target_date})")
                    st.markdown(f"**💬 Ne Yapılacak:** {rec['recommended_quantity']:.0f} adet {rec['symbol']} al, {format_currency(rec['recommended_value']):,.2f} TL harca")
                    # Sadece neden varsa göster
                    if rec.get('action_reason'):
                        st.caption(f"💡 {rec['action_reason']}")
                    
                    # Neden ARTIR dediğinin detaylı özeti
                    st.markdown("**📋 Neden ARTIR Önerisi:**")
                    from dashboard_portfolio_export import generate_increase_reasons
                    increase_reasons = generate_increase_reasons(rec)
                    for reason in increase_reasons:
                        st.markdown(f"  • {reason}")
                    
                    st.divider()
            
            # 4. TUT/BEKLETİLECEK HİSSELER - Her zaman göster, detaylı açıkla
            if hold_actions:
                # Aktif işlem varsa "BEKLE-GÖR" yerine daha açıklayıcı başlık
                if buy_actions or sell_actions:
                    st.markdown("#### 🟡 4️⃣ Takip Edilecek Hisseler (İşlem yok)")
                else:
                    st.markdown("#### 🟡 BEKLE-GÖR - Detaylı Analiz")
                
                # TUT önerilerini detaylı göster - HER ZAMAN AÇIK
                st.markdown("**📊 Pozisyon Analizi:**")
                for rec in hold_actions:
                    # Mevcut pozisyon bilgisi
                    current_value = rec.get('current_value', 0)
                    profit_loss = rec.get('profit_loss', 0)
                    profit_loss_pct = rec.get('profit_loss_pct', 0)
                    
                    # Renk kodlu bilgi
                    if profit_loss_pct > 0:
                        status_emoji = "🟢"
                        status_color = "green"
                    elif profit_loss_pct < -5:
                        status_emoji = "🔴"
                        status_color = "red"
                    else:
                        status_emoji = "🟡"
                        status_color = "orange"
                    
                    # Kar/zarar rengi
                    profit_color = "green" if profit_loss >= 0 else "red"
                    
                    # Model güveni bilgisi
                    confidence_html = ""
                    if rec.get('confidence'):
                        confidence_val = rec.get('confidence', 0) * 100
                        confidence_html = f"<p style='margin: 5px 0; color: #6c757d; font-size: 0.9em;'>🤖 Model Güveni: {confidence_val:.0f}%</p>"
                    
                    # Hedef fiyat bilgisi
                    target_price_html = ""
                    target_price = rec.get('target_price', rec.get('current_price', 0))
                    current_price = rec.get('current_price', 0)
                    if target_price and target_price != current_price and current_price > 0:
                        potential_return = ((target_price - current_price) / current_price) * 100
                        return_color = "green" if potential_return > 0 else "red"
                        target_price_html = f"<p style='margin: 5px 0;'><strong>🎯 Hedef Fiyat:</strong> <span style='color: {return_color}; font-weight: bold;'>{target_price:,.2f} TL</span> <span style='color: {return_color};'>(%{potential_return:+.1f})</span></p>"
                        
                        # Hedef tarih bilgisi
                        target_days = rec.get('target_days', 30)
                        target_min_date = rec.get('target_min_date', '')
                        target_max_date = rec.get('target_max_date', '')
                        if target_min_date and target_max_date:
                            target_price_html += f"<p style='margin: 5px 0; color: #6c757d; font-size: 0.9em;'>📅 Hedef Tarih: {target_min_date} - {target_max_date}</p>"
                        elif target_days:
                            target_date = (datetime.now() + timedelta(days=target_days)).strftime('%d.%m.%Y')
                            target_price_html += f"<p style='margin: 5px 0; color: #6c757d; font-size: 0.9em;'>📅 Tahmini Süre: ~{target_days} gün ({target_date})</p>"
                    
                    st.markdown(f"""
                    <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid {status_color}">
                        <h4 style="margin: 0 0 10px 0;">{status_emoji} <strong>{rec['symbol']}</strong></h4>
                        <p style="margin: 5px 0;"><strong>Mevcut Pozisyon:</strong> {rec['quantity']:.0f} adet × {current_price:.2f} TL = {current_value:,.0f} TL</p>
                        <p style="margin: 5px 0;"><strong>Ortalama Maliyet:</strong> {rec.get('avg_cost', 0):.2f} TL</p>
                        <p style="margin: 5px 0;"><strong>Kar/Zarar:</strong> <span style="color: {profit_color}; font-weight: bold;">{profit_loss:+,.0f} TL ({profit_loss_pct:+.1f}%)</span></p>
                        {target_price_html}
                        <p style="margin: 5px 0; padding-top: 10px; border-top: 1px solid #ddd;"><strong>💡 Analiz:</strong> {rec['action_reason']}</p>
                        {confidence_html}
                    </div>
                    """, unsafe_allow_html=True)
            
            # Toplam işlem özeti
            total_buy = sum([r['recommended_value'] for r in buy_actions])
            total_sell = sum([r['recommended_value'] for r in sell_actions])
            
            # Export butonu ekle
            st.markdown("---")
            st.markdown("### 📄 Rapor Export")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Export için verileri hazırla
                export_date = datetime.now()
                export_portfolio = {
                    'cash': portfolio.get('cash', 0),
                    'stocks': portfolio.get('stocks', {})
                }
                
                try:
                    from dashboard_portfolio_export import create_portfolio_recommendations_export
                    
                    # Session state key
                    export_key = f"export_word_{len(recommendations)}_{export_date.strftime('%Y%m%d')}"
                    
                    # Word dosyasını oluştur
                    if export_key not in st.session_state:
                        if st.button("📝 Word Raporu Oluştur", type="primary", use_container_width=True):
                            with st.spinner("📄 Word raporu oluşturuluyor..."):
                                doc = create_portfolio_recommendations_export(recommendations, export_portfolio, export_date)
                                
                                # Logs klasörünü oluştur
                                logs_dir = "logs"
                                os.makedirs(logs_dir, exist_ok=True)
                                
                                # Dosya adı
                                filename = f"portfoy_onerileri_{export_date.strftime('%Y%m%d_%H%M%S')}.docx"
                                filepath = os.path.join(logs_dir, filename)
                                
                                # Dosyayı kaydet
                                doc.save(filepath)
                                
                                # Dosyayı oku ve session state'e kaydet
                                with open(filepath, "rb") as f:
                                    file_data = f.read()
                                    st.session_state[export_key] = {
                                        'data': file_data,
                                        'filename': filename
                                    }
                                
                                st.success(f"✅ Word raporu oluşturuldu!")
                                st.rerun()
                    
                    # Download butonu
                    if export_key in st.session_state:
                        file_data = st.session_state[export_key]['data']
                        filename = st.session_state[export_key]['filename']
                        
                        st.download_button(
                            label="⬇️ Word Dosyasını İndir",
                            data=file_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                            key=f"download_word_{export_date.strftime('%Y%m%d_%H%M%S')}",
                            type="primary",
                            use_container_width=True
                        )
                        
                        if st.button("🔄 Yeniden Oluştur", use_container_width=True):
                            del st.session_state[export_key]
                            st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Export hatası: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
            
            with col2:
                st.info("""
                **📋 Export Özellikleri:**
                - Tüm alım/satım önerileri
                - Neden AL/ARTIR/SAT dediğinin detayları
                - Portföy özeti ve işlem tutarları
                - Takip edilecek hisseler listesi
                """)
            
            st.markdown("---")
            st.markdown("### 📊 İşlem Özeti")
            
            # İşlem özeti - kompakt format
            current_total = sum([s['quantity'] * s['avg_cost'] for s in portfolio['stocks'].values()])
            net_cash = portfolio['cash'] + total_sell - total_buy
            
            st.markdown("""
            <div style="background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%); 
                        padding: 20px; 
                        border-radius: 10px; 
                        margin: 10px 0;">
                <table style="width: 100%; border-collapse: collapse;">
                    <tr style="border-bottom: 2px solid #ddd;">
                        <td style="padding: 8px; font-weight: bold;">📊 Portföy Değeri</td>
                        <td style="padding: 8px; text-align: right; font-weight: bold; color: #28a745;">{:.2f} TL</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #eee;">
                        <td style="padding: 8px;">💵 Başlangıç Nakit</td>
                        <td style="padding: 8px; text-align: right;">{:.2f} TL</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #eee;">
                        <td style="padding: 8px;">💰 Toplam Satış</td>
                        <td style="padding: 8px; text-align: right; color: #dc3545;">+{:.2f} TL</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #eee;">
                        <td style="padding: 8px;">💸 Toplam Alım</td>
                        <td style="padding: 8px; text-align: right; color: #6c757d;">-{:.2f} TL</td>
                    </tr>
                    <tr style="background: #28a745; color: white; border-top: 2px solid #28a745;">
                        <td style="padding: 10px; font-weight: bold;">💵 Kalan Nakit</td>
                        <td style="padding: 10px; text-align: right; font-weight: bold;">{:.2f} TL</td>
                    </tr>
                </table>
            </div>
            """.format(current_total, portfolio['cash'], total_sell, total_buy, net_cash), unsafe_allow_html=True)
        else:
            st.info("📝 Henüz hisse eklenmemiş veya analiz için yeterli veri yok.")
    else:
        st.info("💡 Lütfen nakit para veya portföy bilgilerinizi girin.")
    
    # === ALIŞ-SATIŞ İŞLEMLERİ KAYIT SİSTEMİ ===
    st.markdown("---")
    st.markdown("### 📝 Alış-Satış İşlemleri")
    
    # Duplicate temizleme butonu (sadece transaction varsa göster)
    transactions = load_user_transactions(user_id)
    
    if transactions:
        # Duplicate kontrolü yap
        db = get_db()
        with st.expander("🔧 Gelişmiş Seçenekler", expanded=False):
            if st.button("🧹 Duplicate Transaction'ları Temizle", type="secondary", help="Aynı işlemlerden tekrarlananları sil"):
                with st.spinner("Duplicate transaction'lar temizleniyor..."):
                    deleted_count = db.remove_duplicate_transactions(user_id)
                    if deleted_count > 0:
                        st.success(f"✅ {deleted_count} duplicate transaction temizlendi!")
                        st.rerun()
                    else:
                        st.info("ℹ️ Duplicate transaction bulunamadı.")
    
    st.info("💡 Alış-satış işlemlerinizi kaydedin ve Excel formatında export edin.")
    
    # İşlem sayısı bilgisi ve başarı mesajı
    if transactions:
        st.success(f"✅ {len(transactions)} işlem otomatik olarak yüklendi!")
        
        # Son işlem bilgisi
        if len(transactions) > 0:
            last_transaction = transactions[0]  # En yeni işlem (tarihe göre sıralı)
            last_date = datetime.fromisoformat(last_transaction.get('date', datetime.now().isoformat())).strftime('%d.%m.%Y')
            st.caption(f"📅 Son işlem: {last_transaction.get('type', '')} - {last_transaction.get('symbol', '')} ({last_date})")
    else:
        st.info("💡 Henüz işlem kaydedilmemiş. Aşağıdaki formu kullanarak ilk işleminizi ekleyin.")
    
    # Form değerlerini session state'te sakla (kullanıcı deneyimi için)
    if 'form_transaction_type' not in st.session_state:
        st.session_state.form_transaction_type = "AL"
    if 'form_transaction_symbol' not in st.session_state:
        st.session_state.form_transaction_symbol = get_all_bist_stocks()[0] if get_all_bist_stocks() else ""
    if 'form_transaction_quantity' not in st.session_state:
        st.session_state.form_transaction_quantity = 100
    if 'form_transaction_price' not in st.session_state:
        st.session_state.form_transaction_price = 100.0
    if 'form_transaction_date' not in st.session_state:
        st.session_state.form_transaction_date = datetime.now().date()
    
    # Alış-Satış Formu
    with st.form("transaction_form", clear_on_submit=False):
        st.markdown("#### ➕ Yeni İşlem Ekle")
        
        col1, col2 = st.columns(2)
        
        with col1:
            transaction_type = st.selectbox(
                "İşlem Tipi:",
                ["AL", "SAT"],
                index=0 if st.session_state.form_transaction_type == "AL" else 1,
                key="transaction_type"
            )
            
            transaction_symbol = st.selectbox(
                "Hisse:",
                get_all_bist_stocks(),
                index=get_all_bist_stocks().index(st.session_state.form_transaction_symbol) if st.session_state.form_transaction_symbol in get_all_bist_stocks() else 0,
                key="transaction_symbol"
            )
            
            transaction_quantity = st.number_input(
                "Adet:",
                min_value=1,
                value=st.session_state.form_transaction_quantity,
                step=10,
                key="transaction_quantity"
            )
        
        with col2:
            transaction_price = st.number_input(
                "Birim Fiyat (TL):",
                min_value=0.01,
                value=st.session_state.form_transaction_price,
                step=0.10,
                format="%.2f",
                key="transaction_price"
            )
            
            transaction_date = st.date_input(
                "İşlem Tarihi:",
                value=st.session_state.form_transaction_date,
                key="transaction_date"
            )
        
        submit_button = st.form_submit_button("💾 İşlemi Kaydet", type="primary", use_container_width=True)
        
        # Form değerlerini her zaman session state'e kaydet (kullanıcı deneyimi için)
        if transaction_type:
            st.session_state.form_transaction_type = transaction_type
        if transaction_symbol:
            st.session_state.form_transaction_symbol = transaction_symbol
        if transaction_quantity:
            st.session_state.form_transaction_quantity = transaction_quantity
        if transaction_price:
            st.session_state.form_transaction_price = transaction_price
        if transaction_date:
            st.session_state.form_transaction_date = transaction_date
        
        if submit_button:
            total_value = transaction_quantity * transaction_price
            
            transaction = {
                'type': transaction_type,
                'symbol': transaction_symbol,
                'quantity': int(transaction_quantity),
                'price': float(transaction_price),
                'total_value': float(total_value),
                'date': transaction_date.isoformat()
            }
            
            try:
                save_user_transaction(user_id, transaction)
                st.success(f"✅ {transaction_type} işlemi kaydedildi: {transaction_quantity} adet {transaction_symbol} @ {transaction_price:.2f} TL")
                
                # Başarılı kayıt sonrası formu temizle (sadece tarih bugünün tarihi olsun)
                st.session_state.form_transaction_date = datetime.now().date()
                # Diğer değerleri koru (kullanıcı aynı hisse için tekrar işlem yapabilir)
                
                # İşlemlerin görünmesi için sayfayı yenile
                st.rerun()
            except Exception as e:
                st.error(f"❌ İşlem kaydedilirken hata oluştu: {str(e)}")
                # Hata durumunda değerler zaten session state'te korunuyor
    
    # İşlemleri Görüntüleme
    if transactions:
        st.markdown("---")
        st.markdown("#### 📊 İşlem Geçmişi")
        
        # Filtreleme
        col1, col2, col3 = st.columns(3)
        
        with col1:
            filter_type = st.selectbox(
                "İşlem Tipi Filtresi:",
                ["Tümü", "AL", "SAT"],
                key="filter_type"
            )
        
        with col2:
            filter_symbol = st.selectbox(
                "Hisse Filtresi:",
                ["Tümü"] + sorted(list(set([t['symbol'] for t in transactions]))),
                key="filter_symbol"
            )
        
        with col3:
            # Tarih filtresi
            date_range = st.selectbox(
                "Tarih Aralığı:",
                ["Tümü", "Son 1 Ay", "Son 3 Ay", "Son 6 Ay", "Son 1 Yıl"],
                key="date_range"
            )
        
        # Filtreleme uygula
        filtered_transactions = transactions.copy()
        
        if filter_type != "Tümü":
            filtered_transactions = [t for t in filtered_transactions if t['type'] == filter_type]
        
        if filter_symbol != "Tümü":
            filtered_transactions = [t for t in filtered_transactions if t['symbol'] == filter_symbol]
        
        if date_range != "Tümü":
            now = datetime.now()
            if date_range == "Son 1 Ay":
                cutoff_date = now - timedelta(days=30)
            elif date_range == "Son 3 Ay":
                cutoff_date = now - timedelta(days=90)
            elif date_range == "Son 6 Ay":
                cutoff_date = now - timedelta(days=180)
            elif date_range == "Son 1 Yıl":
                cutoff_date = now - timedelta(days=365)
            
            filtered_transactions = [
                t for t in filtered_transactions 
                if datetime.fromisoformat(t['date']) >= cutoff_date
            ]
        
        # İşlemleri göster
        if filtered_transactions:
            # İşlemleri tarihe göre sırala (en yeni en üstte)
            sorted_transactions = sorted(
                filtered_transactions, 
                key=lambda x: datetime.fromisoformat(x['date']) if isinstance(x['date'], str) else x['date'],
                reverse=True
            )
            
            # Tablo başlığı
            st.markdown("**İşlem Listesi:**")
            
            # Tablo başlık satırı
            header_cols = st.columns([1, 1, 1, 1, 1, 1, 0.5])
            with header_cols[0]:
                st.markdown("**Tarih**")
            with header_cols[1]:
                st.markdown("**Hisse**")
            with header_cols[2]:
                st.markdown("**İşlem Tipi**")
            with header_cols[3]:
                st.markdown("**Adet**")
            with header_cols[4]:
                st.markdown("**Birim Fiyat**")
            with header_cols[5]:
                st.markdown("**Toplam Tutar**")
            with header_cols[6]:
                st.markdown("**İşlem**")
            
            st.markdown("<hr style='margin: 5px 0;'>", unsafe_allow_html=True)
            
            # Her işlemi bir satır olarak göster
            for idx, transaction in enumerate(sorted_transactions):
                t_id = transaction.get('id')
                if t_id is None:
                    # Eğer ID yoksa, index kullan (geçici çözüm)
                    t_id = idx
                
                # Tarih formatını düzenle
                date_str = transaction['date']
                if isinstance(date_str, str):
                    try:
                        date_obj = datetime.fromisoformat(date_str)
                        date_str = date_obj.strftime('%d.%m.%Y')
                    except:
                        pass
                
                # İşlem tipi için renk
                type_color = "🟢" if transaction['type'] == 'AL' else "🔴"
                type_text = f"{type_color} {transaction['type']}"
                
                # Satır oluştur
                cols = st.columns([1, 1, 1, 1, 1, 1, 0.5])
                
                with cols[0]:
                    st.write(f"**{date_str}**")
                
                with cols[1]:
                    st.write(f"**{transaction['symbol']}**")
                
                with cols[2]:
                    st.write(type_text)
                
                with cols[3]:
                    st.write(f"{transaction['quantity']} adet")
                
                with cols[4]:
                    st.write(f"{transaction['price']:.2f} TL")
                
                with cols[5]:
                    st.write(f"**{transaction['total_value']:,.2f} TL**")
                
                with cols[6]:
                    # Silme butonu
                    if st.button("🗑️", key=f"delete_{t_id}_{idx}", help="İşlemi sil"):
                        if delete_user_transaction(user_id, t_id):
                            st.success(f"✅ İşlem başarıyla silindi!")
                            st.rerun()
                        else:
                            st.error("❌ İşlem silinirken bir hata oluştu.")
                
                # Satırlar arası ayırıcı
                if idx < len(sorted_transactions) - 1:
                    st.markdown("<hr style='margin: 5px 0;'>", unsafe_allow_html=True)
            
            # Kar/Zarar Hesaplama
            st.markdown("---")
            st.markdown("#### 💰 Kar/Zarar Analizi")
            
            profit_loss_data = calculate_profit_loss(filtered_transactions)
            
            # Özet metrikler
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_buy = sum([t['total_value'] for t in filtered_transactions if t['type'] == 'AL'])
                st.metric("📈 Toplam Alış", f"{format_currency(total_buy):,.2f} TL")
            
            with col2:
                total_sell = sum([t['total_value'] for t in filtered_transactions if t['type'] == 'SAT'])
                st.metric("📉 Toplam Satış", f"{format_currency(total_sell):,.2f} TL")
            
            with col3:
                total_profit_loss = profit_loss_data.get('total_profit_loss', 0)
                profit_color = "normal" if total_profit_loss >= 0 else "inverse"
                st.metric("💰 Toplam Kar/Zarar", f"{format_currency(total_profit_loss):+,.2f} TL", delta=None)
            
            with col4:
                st.metric("📊 Toplam İşlem", f"{len(filtered_transactions)}")
            
            # Hisse bazında kar/zarar
            symbol_profits = profit_loss_data.get('symbol_profits', {})
            if symbol_profits:
                st.markdown("**Hisse Bazında Kar/Zarar:**")
                profit_df = pd.DataFrame([
                    {
                        'Hisse': symbol,
                        'Toplam Kar/Zarar (TL)': data['profit_loss'],
                        'Kar/Zarar (%)': f"{data['profit_loss_pct']:.2f}%",
                        'İşlem Sayısı': data['count']
                    }
                    for symbol, data in symbol_profits.items()
                ])
                st.dataframe(profit_df, use_container_width=True, hide_index=True)
            
            # Excel/CSV Export
            st.markdown("---")
            st.markdown("#### 📄 İşlem Listesini Dışa Aktar")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Excel Export
                if OPENPYXL_AVAILABLE:
                    try:
                        excel_data = export_transactions_to_excel(filtered_transactions, profit_loss_data)
                        st.download_button(
                            label="📥 Excel Dosyasını İndir (.xlsx)",
                            data=excel_data,
                            file_name=f"alis_satis_islemleri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error(f"❌ Excel export hatası: {str(e)}")
                else:
                    st.warning("⚠️ Excel export için `openpyxl` modülü gerekli. CSV formatını kullanabilirsiniz.")
                    st.info("💡 Excel export için: `pip install openpyxl` komutunu çalıştırın.")
                
                # CSV Export (her zaman mevcut)
                try:
                    csv_data = export_transactions_to_csv(filtered_transactions)
                    st.download_button(
                        label="📥 CSV Dosyasını İndir (.csv)",
                        data=csv_data,
                        file_name=f"alis_satis_islemleri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                        type="secondary",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"❌ CSV export hatası: {str(e)}")
            
            with col2:
                if OPENPYXL_AVAILABLE:
                    st.info("""
                    **Excel Dosyası İçeriği:**
                    - 📋 İşlemler sayfası: Tüm alış-satış işlemleri
                    - 💰 Kar/Zarar Analizi sayfası: 
                      • Özet metrikler (Toplam Alış, Satış, Kar/Zarar)
                      • Hisse bazında detaylı kar/zarar analizi
                    - 📊 Özet sayfası: Toplam metrikler
                    """)
                else:
                    st.info("""
                    **CSV Dosyası İçeriği:**
                    - 📋 Tüm alış-satış işlemleri
                    - 📊 Excel'de açılabilir format
                    - 💾 Hafif ve hızlı
                    """)
        else:
            st.info("🔍 Seçilen filtreler için işlem bulunamadı.")
    else:
        st.info("📝 Henüz işlem kaydedilmemiş. Yukarıdaki formu kullanarak ilk işleminizi ekleyin.")

